import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from collections import OrderedDict, defaultdict
from datetime import datetime


def month_key(date_value):
    """Return key for month grouping as 'YYYY-MM MonthName'."""
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m %B')
    try:
        dt = datetime.fromisoformat(str(date_value))
        return dt.strftime('%Y-%m %B')
    except Exception:
        return str(date_value)


def birim_fiyat_analizi(wb_path: str, veri_sheet: str = 'Sheet1'):
    """Replicate the VBA BirimFiyatAnalizi procedure using openpyxl."""
    wb = openpyxl.load_workbook(wb_path)
    ws_veri = wb[veri_sheet]

    if 'Birim Fiyat Analizi' in wb.sheetnames:
        ws_fiyat = wb['Birim Fiyat Analizi']
        wb.remove(ws_fiyat)
        ws_fiyat = wb.create_sheet('Birim Fiyat Analizi')
    else:
        ws_fiyat = wb.create_sheet('Birim Fiyat Analizi')

    ws_fiyat['A1'] = 'BİRİM FİYAT ANALİZİ - STOK DEVİR SİSTEMİ'
    ws_fiyat.merge_cells('A1:G1')
    hdr = ws_fiyat['A1']
    hdr.font = Font(size=16, bold=True, color='FFFFFF')
    hdr.alignment = Alignment(horizontal='center')
    hdr.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    fish_types = OrderedDict()
    for row in ws_veri.iter_rows(min_row=2, values_only=True):
        fish = row[3]
        if fish:
            fish_types.setdefault(fish, None)

    row_idx = 3
    for fish in fish_types:
        cell = ws_fiyat.cell(row=row_idx, column=1, value=fish.upper())
        ws_fiyat.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        cell.font = Font(size=16, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        row_idx += 1

        row_idx = balik_devir_analizi(ws_veri, ws_fiyat, fish, row_idx)
        row_idx += 1

    birim_fiyat_formatla(ws_fiyat)
    wb.save(wb_path)


def balik_devir_analizi(ws_veri, ws_fiyat, fish_type: str, row_idx: int) -> int:
    months = OrderedDict()
    for row in ws_veri.iter_rows(min_row=2, values_only=True):
        tarih = row[0]
        fish = row[3]
        if fish == fish_type:
            key = month_key(tarih)
            months.setdefault(key, tarih)

    sorted_months = sorted(months.items(), key=lambda x: months[x[0]])

    devir_kg = defaultdict(float)
    devir_kasa = defaultdict(float)

    for key, _ in sorted_months:
        cell = ws_fiyat.cell(row=row_idx, column=1, value=key.upper())
        ws_fiyat.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=7)
        cell.font = Font(size=14, bold=True)
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        row_idx += 1

        headers = [
            'BİRİM FİYAT', 'ÖNCEKİ DEVİR', 'GİRİŞ',
            'ÇIKIŞ', 'NET DURUM', 'KALAN STOK', 'SONRAKİ DEVİR'
        ]
        for col, title in enumerate(headers, 1):
            ws_fiyat.cell(row=row_idx, column=col, value=title)
            ws_fiyat.cell(row=row_idx, column=col).font = Font(bold=True)
            ws_fiyat.cell(row=row_idx, column=col).fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
        row_idx += 1

        row_idx = aylik_devir_analizi(ws_veri, ws_fiyat, key, fish_type, row_idx, devir_kg, devir_kasa)
        row_idx += 1
    return row_idx


def aylik_devir_analizi(ws_veri, ws_fiyat, month_name: str, fish_type: str, row_idx: int, devir_kg: dict, devir_kasa: dict) -> int:
    fiyatlar = OrderedDict()
    for row in ws_veri.iter_rows(min_row=2, values_only=True):
        tarih, _, _, fish, kasa, kg, fiyat = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        if fish == fish_type and month_key(tarih) == month_name:
            if fiyat is not None:
                fiyatlar.setdefault(float(fiyat), None)

    for f in list(devir_kg.keys()):
        if devir_kg[f] > 0 or devir_kasa.get(f, 0) > 0:
            fiyatlar.setdefault(float(f), None)

    for price in fiyatlar:
        prev_kg = devir_kg.get(price, 0.0)
        prev_kasa = devir_kasa.get(price, 0.0)
        giris_kg = giris_kasa = cikis_kg = cikis_kasa = 0.0

        for row in ws_veri.iter_rows(min_row=2, values_only=True):
            tarih, _, _, fish, kasa, kg, fyt = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
            if fish == fish_type and month_key(tarih) == month_name and float(fyt) == price:
                if kasa is not None:
                    if kasa > 0:
                        giris_kasa += kasa
                    else:
                        cikis_kasa += abs(kasa)
                if kg is not None:
                    if kg > 0:
                        giris_kg += kg
                    else:
                        cikis_kg += abs(kg)

        kalan_kg = prev_kg + giris_kg - cikis_kg
        kalan_kasa = prev_kasa + giris_kasa - cikis_kasa
        next_kg = kalan_kg if kalan_kg > 0 else 0
        next_kasa = kalan_kasa if kalan_kasa > 0 else 0
        devir_kg[price] = next_kg
        devir_kasa[price] = next_kasa

        if any([giris_kg, giris_kasa, cikis_kg, cikis_kasa, prev_kg, prev_kasa]):
            ws_fiyat.cell(row=row_idx, column=1, value=f"{price:.2f}")
            ws_fiyat.cell(row=row_idx, column=2, value=f"KG:{prev_kg:.2f} KASA:{prev_kasa:.2f}")
            ws_fiyat.cell(row=row_idx, column=3, value=f"KG:{giris_kg:.2f} KASA:{giris_kasa:.2f}")
            ws_fiyat.cell(row=row_idx, column=4, value=f"KG:{cikis_kg:.2f} KASA:{cikis_kasa:.2f}")
            ws_fiyat.cell(row=row_idx, column=5, value=f"KG:{kalan_kg:.2f} KASA:{kalan_kasa:.2f}")
            ws_fiyat.cell(row=row_idx, column=6, value=f"KG:{next_kg:.2f} KASA:{next_kasa:.2f}")
            ws_fiyat.cell(row=row_idx, column=7, value=f"KG:{next_kg:.2f} KASA:{next_kasa:.2f}")

            net_val = kalan_kg + kalan_kasa
            color = '92D050' if net_val > 0 else 'FFC000' if net_val == 0 else 'FF6347'
            for col in (5, 6, 7):
                ws_fiyat.cell(row=row_idx, column=col).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            row_idx += 1
    return row_idx


def birim_fiyat_formatla(ws_fiyat):
    ws_fiyat.column_dimensions['A'].width = 15
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws_fiyat.column_dimensions[col].width = 20
    for row in ws_fiyat.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(name='Arial', size=10)
    thin = Side(border_style='thin', color='000000')
    max_row = ws_fiyat.max_row
    for row in ws_fiyat.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in range(1, max_row + 1):
        ws_fiyat.row_dimensions[row].height = 20
